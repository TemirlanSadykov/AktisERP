from collections import defaultdict
import json
import openpyxl # type: ignore

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.cache.backends.base import DEFAULT_TIMEOUT
from django.db import transaction
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy, reverse
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView 
from django.views import View
from django.db.models import Q
from django.db.models import IntegerField
from decimal import Decimal
from django.db.models.functions import Cast
from django.db.models import Prefetch
from django.db.models import Prefetch, Sum, Value, DecimalField
from django.db.models.functions import Coalesce, Cast
from django.views.decorators.clickjacking import xframe_options_exempt

from ..decorators import technologist_required
from ..forms import *
from ..models import *


CACHE_TTL = getattr(settings, 'CACHE_TTL', DEFAULT_TIMEOUT)

# @cache_page(CACHE_TTL)
@login_required
@technologist_required
def technologist_page(request):
    context = {
               'sidebar_type': 'technology'
               }
    return render(request, 'technologist_page.html', context)
    
@method_decorator([login_required, technologist_required], name='dispatch')
class ClientListView(ListView):
    model = Client
    template_name = 'technologist/clients/list.html'
    context_object_name = 'clients'
    paginate_by = 10
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    def get_queryset(self):
        return Client.objects.filter(
            is_archived=False
            ).order_by('name')

@method_decorator([login_required, technologist_required], name='dispatch')
class ArchivedClientListView(ListView):
    template_name = 'technologist/clients/list.html'
    context_object_name = 'clients'
    paginate_by = 10
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    def get_queryset(self):
        return Client.objects.filter(            
            is_archived=True
            ).order_by('name')

@method_decorator([login_required, technologist_required], name='dispatch')
class ClientCreateView(CreateView):
    model = Client
    form_class = ClientForm
    template_name = 'technologist/clients/create.html'
    success_url = reverse_lazy('client_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class ClientDetailView(DetailView):
    model = Client
    template_name = 'technologist/clients/detail.html'
    context_object_name = 'client'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class ClientUpdateView(UpdateView):
    model = Client
    form_class = ClientForm
    template_name = 'technologist/clients/edit.html'
    def get_success_url(self):
        return reverse('client_detail', kwargs={'pk': self.object.pk})
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class ClientDeleteView(DeleteView):
    model = Client
    template_name = 'technologist/clients/delete.html'
    success_url = reverse_lazy('client_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class ClientArchiveView(UpdateView):
    model = Client
    template_name = 'technologist/clients/delete.html'
    success_url = reverse_lazy('client_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    def post(self, request, *args, **kwargs):
        clien_order = self.get_object()
        clien_order.is_archived = True
        clien_order.save()
        return HttpResponseRedirect(self.success_url)


@method_decorator([login_required, technologist_required], name='dispatch')
class ClientUnArchiveView(UpdateView):
    model = Client
    template_name = 'technologist/clients/delete.html'
    success_url = reverse_lazy('client_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    def post(self, request, *args, **kwargs):
        clien_order = self.get_object()
        clien_order.is_archived = False
        clien_order.save()
        return HttpResponseRedirect(self.success_url)



@method_decorator([login_required, technologist_required], name='dispatch')
class ClientOrderListView(ListView):
    model = ClientOrder
    template_name = 'technologist/client/orders/list.html'
    context_object_name = 'orders'
    paginate_by = 10
    form_class = DateRangeForm 

    def get_queryset(self):
        queryset = super().get_queryset().filter(is_archived=False)
        today = timezone.localdate()

        # Read the term filter from GET. Defaults to 'upcoming' if not provided.
        term_filter = self.request.GET.get('term', 'upcoming').lower()

        if term_filter == 'upcoming':
            # Upcoming orders: term is today or later; sort ascending (soonest first)
            queryset = queryset.filter(term__gte=today).order_by('term')
        elif term_filter == 'passed':
            # Passed orders: term is before today; sort descending (most recent first)
            queryset = queryset.filter(term__lt=today).order_by('-term')
        else:
            # Default to upcoming if unknown value is provided
            queryset = queryset.filter(term__gte=today).order_by('term')

        # Apply optional date range filtering
        form = self.form_class(self.request.GET)
        if form.is_valid():
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')
            if start_date and end_date:
                queryset = queryset.filter(launch__range=[start_date, end_date])
            elif start_date:
                queryset = queryset.filter(launch__gte=start_date)
            elif end_date:
                queryset = queryset.filter(launch__lte=end_date)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET or None)
        today = timezone.localdate()
        orders_with_days_left = []

        # Calculate days_left for display. (Negative values for passed orders.)
        for order in context['orders']:
            days_left = (order.term - today).days
            orders_with_days_left.append({'order': order, 'days_left': days_left})
        context['orders_with_days_left'] = orders_with_days_left

        # Pass the current term filter for use in the template
        context['term_filter'] = self.request.GET.get('term', 'upcoming').lower()
        context['ClientOrder'] = ClientOrder
        context['sidebar_type'] = 'technology'
        return context
    
@method_decorator([login_required, technologist_required], name='dispatch')
class ClientOrderCreateView(CreateView):
    model = ClientOrder
    form_class = ClientOrderForm
    template_name = 'technologist/client/orders/create.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        form.save_m2m()
        redirect_url = reverse('client_order_detail', kwargs={'pk': self.object.pk})
        return HttpResponseRedirect(redirect_url)
    
@method_decorator([login_required, technologist_required], name='dispatch')
class ClientOrderDetailView(DetailView):
    model = ClientOrder
    form_class = ClientOrderForm
    template_name = 'technologist/client/orders/detail.html'
    context_object_name = 'client_order'

    def get_context_data(self, **kwargs):
        context = super(ClientOrderDetailView, self).get_context_data(**kwargs)
        client_order = context['client_order']
        context['orders'] = client_order.orders.all()
        today = timezone.localdate()
        if client_order.term >= today:
            days_left = (client_order.term - today).days
        else:
            days_left = 0
        context['days_left'] = days_left
        context['sidebar_type'] = 'technology'
        return context
    
@method_decorator([login_required, technologist_required], name='dispatch')
class ClientOrderUpdateView(UpdateView):
    model = ClientOrder
    form_class = ClientOrderForm
    template_name = 'technologist/client/orders/edit.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

    def get_success_url(self):
        return reverse('client_order_detail', kwargs={'pk': self.object.pk})
    
@method_decorator([login_required, technologist_required], name='dispatch')
class ClientOrderDeleteView(DeleteView):
    model = ClientOrder
    template_name = 'technologist/client/orders/delete.html'
    success_url = reverse_lazy('client_order_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class ClientOrderArchiveView(UpdateView):
    model = ClientOrder
    template_name = 'technologist/client/orders/delete.html'
    success_url = reverse_lazy('client_order_list')
        
    def post(self, request, *args, **kwargs):
        clien_order = self.get_object()
        clien_order.is_archived = True
        clien_order.save()
        return HttpResponseRedirect(self.success_url)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    

@method_decorator([login_required, technologist_required], name='dispatch')
class ClientOrderUnArchiveView(UpdateView):
    model = ClientOrder
    template_name = 'technologist/client/orders/delete.html'
    success_url = reverse_lazy('client_order_list')
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    def post(self, request, *args, **kwargs):
        clien_order = self.get_object()
        clien_order.is_archived = False
        clien_order.save()
        return HttpResponseRedirect(self.success_url)
    
@method_decorator([login_required, technologist_required], name='dispatch')
class ArchivedClientOrderListView(ListView):
    model = ClientOrder
    template_name = 'technologist/client/orders/list.html'  # Use a separate template if necessary
    context_object_name = 'archived_orders'
    paginate_by = 10
    form_class = DateRangeForm 

    def get_queryset(self):
        queryset = super().get_queryset().filter(is_archived=True)
        today = timezone.localdate()

        # Read the term filter from GET. Defaults to 'upcoming' if not provided.
        term_filter = self.request.GET.get('term', 'upcoming').lower()

        if term_filter == 'upcoming':
            # Upcoming orders: term is today or later; sort ascending (soonest first)
            queryset = queryset.filter(term__gte=today).order_by('term')
        elif term_filter == 'passed':
            # Passed orders: term is before today; sort descending (most recent first)
            queryset = queryset.filter(term__lt=today).order_by('-term')
        else:
            # Default to upcoming if unknown value is provided
            queryset = queryset.filter(term__gte=today).order_by('term')

        # Apply optional date range filtering
        form = self.form_class(self.request.GET)
        if form.is_valid():
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')
            if start_date and end_date:
                queryset = queryset.filter(launch__range=[start_date, end_date])
            elif start_date:
                queryset = queryset.filter(launch__gte=start_date)
            elif end_date:
                queryset = queryset.filter(launch__lte=end_date)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET or None)
        today = timezone.localdate()

        # Calculate days_left for display. (Negative values for passed orders.)
        orders_with_days_left = []
        for order in context['archived_orders']:
            days_left = (order.term - today).days
            orders_with_days_left.append({'order': order, 'days_left': days_left})
        context['orders_with_days_left'] = orders_with_days_left

        # Pass the current term filter for use in the template
        context['term_filter'] = self.request.GET.get('term', 'upcoming').lower()
        context['ClientOrder'] = ClientOrder
        context['sidebar_type'] = 'technology'
        return context
    
@login_required
@technologist_required
@require_POST
def client_order_complete(request, pk):
    client_order = get_object_or_404(ClientOrder, pk=pk)
    STATUSES = [ClientOrder.NEW, ClientOrder.IN_PROGRESS, ClientOrder.COMPLETED]

    # Get the index of the current status
    current_index = STATUSES.index(client_order.status)

    # Calculate the next index, cycling back to 0 if at the end
    next_index = (current_index + 1) % len(STATUSES)

    # Set the status to the next one
    client_order.status = STATUSES[next_index]

    client_order.save()

    return redirect('client_order_detail', pk=client_order.pk)

@require_POST
@login_required
def add_client_api(request):
    form = ClientForm(request.POST)
    if form.is_valid():
        client = form.save()
        data = {
            'success': True,
            'client_id': client.id,
            'client_name': client.name,
            'client_description': client.description,
        }
        return JsonResponse(data)
    else:
        # Return form errors as JSON (status code 400)
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)



@method_decorator([login_required, technologist_required], name='dispatch')
class OrderListView(ListView):
    model = Order
    template_name = 'technologist/orders/list.html'
    context_object_name = 'orders'
    paginate_by = 10

    def get_queryset(self):
        status = self.request.GET.get('status', None)
        search_query = self.request.GET.get('search', None)
        queryset = super().get_queryset().order_by('client_order__term')

        queryset = queryset.filter(client_order__is_archived=False)

        if status:
            try:
                status = int(status)
                if status in dict(self.model.TYPE_CHOICES):
                    queryset = queryset.filter(status=status)
            except ValueError:
                pass

        if search_query:
            queryset = queryset.filter(
                Q(model__name__icontains=search_query) |
                Q(color__icontains=search_query) |
                Q(fabrics__icontains=search_query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.localdate()
        orders_with_days_left = []

        for order in context['orders']:
            days_left = (order.client_order.term - today).days
            orders_with_days_left.append({
                'order': order,
                'days_left': days_left
            })

        orders_with_days_left_sorted = sorted(orders_with_days_left, key=lambda x: x['days_left'])

        context['orders_with_days_left'] = orders_with_days_left_sorted
        context['selected_status'] = self.request.GET.get('status', '')
        context['search_query'] = self.request.GET.get('search', '')
        context['Order'] = Order
        context['sidebar_type'] = 'technology'
        return context
    
@method_decorator([login_required, technologist_required], name='dispatch')
class OrderDetailView(DetailView):
    model = Order
    template_name = 'technologist/orders/detail.html'
    context_object_name = 'order'

    def get_queryset(self):
        return Order.objects.prefetch_related(
            # Prefetch order's size quantities with related color/fabrics.
            Prefetch(
                'size_quantities',
                queryset=SizeQuantity.objects.select_related('color', 'fabrics').order_by('size')
            ),
            # Prefetch cuts, annotate total_layers from passports, and prefetch their related objects.
            Prefetch(
                'cuts',
                queryset=Cut.objects.order_by('number')
                .annotate(total_layers=Coalesce(Sum('passports__layers'), Value(0), output_field=IntegerField()))
                .prefetch_related(
                    'size_quantities',
                    Prefetch(
                        'passports',
                        queryset=Passport.objects.prefetch_related(
                            Prefetch(
                                'size_quantities',
                                queryset=SizeQuantity.objects.select_related('color', 'fabrics')
                            )
                        )
                    )
                )
            ),
            'client_order'  # For days_left calculation.
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        order = context['order']

        # Build pivot data for "Required Quantities" table.
        pivot_data = {}
        all_sizes_set = set()
        for sq in order.size_quantities.all():
            all_sizes_set.add(sq.size)
            key = (sq.color, sq.fabrics)  # relies on __str__ of these models.
            pivot_data.setdefault(key, {})[sq.size] = sq.quantity

        try:
            all_sizes = sorted(all_sizes_set, key=lambda s: int(s))
        except (ValueError, TypeError):
            all_sizes = sorted(all_sizes_set)

        # Associated cuts are already prefetched.
        associated_cuts = order.cuts.all()

        # Calculate days left until the term.
        today = timezone.now().date()
        days_left = (order.client_order.term - today).days if order.client_order.term >= today else 0

        # --- Build detailed cut data ---
        cut_details = []
        for cut in associated_cuts:
            # Get unique sizes used in this cut.
            cut_sizes = list(cut.size_quantities.values_list('size', flat=True).distinct())
            try:
                cut_sizes_sorted = sorted(cut_sizes, key=lambda s: int(s))
            except (ValueError, TypeError):
                cut_sizes_sorted = sorted(cut_sizes)
            
            # Build passport data and aggregate colors from all passports.
            passport_list = []
            aggregated_colors = set()
            for passport in cut.passports.all():
                colors = {psq.color.name for psq in passport.size_quantities.all() if psq.color}
                fabrics = {psq.fabrics.name for psq in passport.size_quantities.all() if psq.fabrics}
                aggregated_colors.update(colors)
                passport_list.append({
                    'passport_id': passport.id,
                    'passport_number': passport.number,
                    'colors': ", ".join(sorted(colors)),
                    'fabrics': ", ".join(sorted(fabrics)),
                })
            aggregated_colors_str = ", ".join(sorted(aggregated_colors))
            
            # Use the annotated total_layers.
            total_layers = cut.total_layers

            cut_details.append({
                'cut_id': cut.id,
                'cut_number': cut.number,
                'cut_date': cut.date,
                'cut_sizes': cut_sizes_sorted,
                'aggregated_colors': aggregated_colors_str,
                'total_layers': total_layers,
                'passports': passport_list,
            })

        context.update({
            'pivot_data': pivot_data,   # Pivoted required quantities.
            'all_sizes': all_sizes,     # Sorted list of sizes for header.
            'first_sq': order.size_quantities.order_by('id').first(),  # For BOM link.
            'days_left': days_left,
            'associated_cuts': associated_cuts,
            'cut_details': cut_details, # Detailed cut info including aggregated colors, sizes, and total layers.
            'sidebar_type': 'technology'
        })
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class OrderCreateView(CreateView):
    model = Order
    form_class = OrderForm
    template_name = 'technologist/orders/create.html'

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.client_order = get_object_or_404(ClientOrder, pk=self.kwargs['client_order_pk'])
        self.object.save()
        form.save_m2m()
        try:
            # create all SizeQuantity instances
            self.handle_size_quantities(self.object, self.request.POST)
        except ValidationError as e:
            self.object.delete()
            form.add_error(None, e.message)
            return self.form_invalid(form)

        # grab the first SizeQuantity you just created
        first_sq = self.object.size_quantities.first()
        if not first_sq:
            # fallback if somehow none were created
            return redirect('order_detail', pk=self.object.pk)

        return redirect('bom_create', pk=first_sq.pk)

    def handle_size_quantities(self, order, post_data):
        """
        Process the dynamic table data:
        - Ensures that each row's (color, fabric) combination is unique.
        - Sums up all provided quantities per cell.
        - Creates SizeQuantity objects (with SKU) and associates them with the order.
        - Updates the order's overall quantity based on the sum.
        """
        used_combinations = set()  # To track (color_id, fabric_id) per row
        used_colors = set()
        used_fabrics = set()
        sizes = []
        total_quantity = 0  # Running total of quantities

        # Retrieve the sizes from header inputs.
        col = 0
        while f'header_{col}' in post_data:
            sizes.append(post_data[f'header_{col}'])
            col += 1

        row = 0
        while f'row-{row}_color' in post_data:
            color_id = post_data.get(f'row-{row}_color')
            fabric_id = post_data.get(f'row-{row}_fabric')

            # Validate uniqueness of the color–fabric combination per row.
            combination_key = (color_id, fabric_id)
            if combination_key in used_combinations:
                raise ValidationError("Цвет-ткань должна быть уникальной.")
            used_combinations.add(combination_key)

            # Lookup the color and fabric objects.
            color = Color.objects.filter(pk=color_id).first() if color_id else None
            fabric = Fabrics.objects.filter(pk=fabric_id).first() if fabric_id else None
            model = order.model
            if color:
                used_colors.add(color)
            if fabric:
                used_fabrics.add(fabric)

            # Process each size column for the current row.
            for col_index, size in enumerate(sizes):
                quantity = post_data.get(f'cell_{row}_{col_index}', 0)
                if quantity and int(quantity) > 0:
                    quantity_int = int(quantity)
                    total_quantity += quantity_int
                    # Get the SKU value; if SKU field is absent or empty, default to an empty string.
                    sku = post_data.get(f'sku_{row}_{col_index}', '').strip()
                    size_qty = SizeQuantity.objects.create(
                        size=size,
                        quantity=quantity_int,
                        factual=quantity_int,
                        model=model,
                        sku=sku,  # Save SKU along with quantity.
                        color=color,
                        fabrics=fabric
                    )
                    order.size_quantities.add(size_qty)
            row += 1

        # Update the order's overall quantity with the calculated total.
        order.quantity = total_quantity
        order.save(update_fields=['quantity'])

        # Save the collected colors and fabrics to the order.
        if used_colors:
            order.colors.add(*used_colors)
        if used_fabrics:
            order.fabrics.add(*used_fabrics)

    def get_context_data(self, **kwargs):
        context = super(OrderCreateView, self).get_context_data(**kwargs)
        context['client_order_pk'] = self.kwargs.get('client_order_pk')
        context['sidebar_type'] = 'technology'
        # Pass available colors and fabrics (and sizes) for use in the template.
        context['colors'] = Color.objects.filter(is_archived=False).order_by('name')
        context['fabrics'] = Fabrics.objects.filter(is_archived=False).order_by('name')
        context['models'] = Model.objects.filter(is_archived=False).order_by('name')
        return context
    
@require_POST
@login_required
def add_color_api(request):
    form = ColorForm(request.POST)
    if form.is_valid():
        color = form.save()
        data = {
            'success': True,
            'color_id': color.id,
            'color_name': color.name,
        }
        return JsonResponse(data)
    else:
        # Return form errors as JSON (status code 400)
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    
@require_POST
@login_required
def add_fabric_api(request):
    form = FabricsForm(request.POST)
    if form.is_valid():
        fabric = form.save()
        data = {
            'success': True,
            'fabric_id': fabric.id,
            'fabric_name': fabric.name,
        }
        return JsonResponse(data)
    else:
        # Return form errors as JSON (status code 400)
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    
@method_decorator([login_required, technologist_required], name='dispatch')
class OrderUpdateView(UpdateView):
    model = Order
    form_class = OrderForm
    template_name = 'technologist/orders/edit.html'

    def get_context_data(self, **kwargs):
        context = super(OrderUpdateView, self).get_context_data(**kwargs)
        context['client_order_pk'] = self.object.client_order.pk
        context['sidebar_type'] = 'technology'
        context['colors'] = Color.objects.filter(is_archived=False).order_by('name')
        context['fabrics'] = Fabrics.objects.filter(is_archived=False).order_by('name')
        
        # --- Build the table header & rows from existing size quantities ---
        size_quantities = self.object.size_quantities.all()
        sizes = []
        for sq in size_quantities:
            if sq.size not in sizes:
                sizes.append(sq.size)
        if not sizes:
            sizes = ['']
        
        rows_dict = {}
        for sq in size_quantities:
            key = (sq.color_id, sq.fabrics_id)
            if key not in rows_dict:
                rows_dict[key] = {}
            # Include sku in the cell data.
            rows_dict[key][sq.size] = {
                'quantity': sq.quantity,
                'id': sq.pk,
                'sku': sq.sku  # Added SKU here.
            }
        
        table_rows = []
        for (color_id, fabric_id), cells in rows_dict.items():
            row_data = {
                'color_id': color_id,
                'fabric_id': fabric_id,
                'cells': []
            }
            for size in sizes:
                if size in cells:
                    row_data['cells'].append({
                        'size': size,
                        'quantity': cells[size]['quantity'],
                        'id': cells[size]['id'],
                        'sku': cells[size]['sku']  # Pass along any saved SKU.
                    })
                else:
                    row_data['cells'].append({
                        'size': size,
                        'quantity': '',
                        'id': '',
                        'sku': ''
                    })
            table_rows.append(row_data)
        
        if not table_rows:
            table_rows = [{
                'color_id': '',
                'fabric_id': '',
                'cells': [{'size': sizes[0], 'quantity': '', 'id': '', 'sku': ''}]
            }]
        
        context['table_sizes'] = sizes
        context['table_rows'] = table_rows
        context['row_count'] = len(table_rows)
        context['col_count'] = len(sizes)
        return context

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        form.save_m2m()
        try:
            self.handle_size_quantities_update(self.object, self.request.POST)
        except ValidationError as e:
            form.add_error(None, e.message)
            return self.form_invalid(form)
        first_sq = self.object.size_quantities.first()
        if not first_sq:
            # fallback if somehow none were created
            return redirect('order_detail', pk=self.object.pk)
        return redirect('bom_create', pk=first_sq.pk)
    def handle_size_quantities_update(self, order, post_data):
        """
        Process the dynamic table for editing:
          • Ensures that each (color, fabric) combination is unique.
          • Updates existing SizeQuantity objects, creates new ones as needed,
            or deletes ones where quantity is now 0/empty.
          • Sums up all provided quantities and updates order.quantity accordingly.
          • Updates the order's many-to-many colors and fabrics.
          • Also handles the SKU value for each cell.
        """
        used_combinations = set()  # Track (color_id, fabric_id) per row for uniqueness.
        used_colors = set()
        used_fabrics = set()
        sizes = []
        total_quantity = 0  # Running total of all quantities

        col = 0
        while f'header_{col}' in post_data:
            sizes.append(post_data[f'header_{col}'])
            col += 1

        processed_sq_ids = set()
        row = 0
        while f'row-{row}_color' in post_data:
            color_id = post_data.get(f'row-{row}_color')
            fabric_id = post_data.get(f'row-{row}_fabric')
            combination_key = (color_id, fabric_id)
            if combination_key in used_combinations:
                raise ValidationError("Цвет-ткань должна быть уникальной.")
            used_combinations.add(combination_key)

            # Lookup color and fabric objects.
            color = Color.objects.filter(pk=color_id).first() if color_id else None
            fabric = Fabrics.objects.filter(pk=fabric_id).first() if fabric_id else None

            if color:
                used_colors.add(color)
            if fabric:
                used_fabrics.add(fabric)
            
            # Process each cell in this row.
            for col_index, size in enumerate(sizes):
                cell_val = post_data.get(f'cell_{row}_{col_index}', '')
                cell_sq_id = post_data.get(f'cell_id_{row}_{col_index}', None)
                # Retrieve SKU value (if any) and strip whitespace.
                sku_val = post_data.get(f'sku_{row}_{col_index}', '').strip()
                if cell_val and cell_val.isdigit() and int(cell_val) > 0:
                    quantity_val = int(cell_val)
                    total_quantity += quantity_val
                    if cell_sq_id:
                        try:
                            sq_obj = SizeQuantity.objects.get(pk=cell_sq_id)
                            sq_obj.quantity = quantity_val
                            sq_obj.factual = quantity_val
                            sq_obj.size = size  # Update header if changed.
                            sq_obj.color = color
                            sq_obj.fabrics = fabric
                            sq_obj.sku = sku_val  # Update SKU.
                            sq_obj.save()
                            processed_sq_ids.add(sq_obj.pk)
                        except SizeQuantity.DoesNotExist:
                            sq_obj = SizeQuantity.objects.create(
                                size=size,
                                quantity=quantity_val,
                                factual=quantity_val,
                                sku=sku_val,
                                color=color,
                                fabrics=fabric
                            )
                            order.size_quantities.add(sq_obj)
                            processed_sq_ids.add(sq_obj.pk)
                    else:
                        sq_obj = SizeQuantity.objects.create(
                            size=size,
                            quantity=quantity_val,
                            sku=sku_val,
                            color=color,
                            fabrics=fabric
                        )
                        order.size_quantities.add(sq_obj)
                        processed_sq_ids.add(sq_obj.pk)
                else:
                    # If cell is empty or 0 but an object existed, remove it.
                    if cell_sq_id:
                        try:
                            sq_obj = SizeQuantity.objects.get(pk=cell_sq_id)
                            order.size_quantities.remove(sq_obj)
                            sq_obj.delete()
                        except SizeQuantity.DoesNotExist:
                            pass
            row += 1

        # (Optional) Remove any SizeQuantity not present in the submitted data.
        for sq in order.size_quantities.all():
            if sq.pk not in processed_sq_ids:
                order.size_quantities.remove(sq)
                sq.delete()

        # Update order many-to-many fields.
        if used_colors:
            order.colors.set(list(used_colors))
        else:
            order.colors.clear()
        if used_fabrics:
            order.fabrics.set(list(used_fabrics))
        else:
            order.fabrics.clear()

        # Finally, update the order's overall quantity.
        order.quantity = total_quantity
        order.save(update_fields=['quantity'])

@method_decorator([login_required, technologist_required], name='dispatch')
class OrderDeleteView(DeleteView):
    model = Order
    template_name = 'technologist/orders/delete.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    def get_success_url(self):
        return reverse('client_order_detail', kwargs={'pk': self.object.client_order.pk})
    
@method_decorator([login_required, technologist_required], name='dispatch')
class CutDetailTechnologistView(DetailView):
    model = Cut
    template_name = 'technologist/cuts/detail.html'
    context_object_name = 'cut'

    def get_context_data(self, **kwargs):
        from collections import defaultdict
        context = super().get_context_data(**kwargs)
        cut_pk = self.kwargs.get('pk')
        cut = get_object_or_404(Cut, pk=cut_pk)

        # Get all passports related to the cut.
        passports = cut.passports.all().order_by('-number')

        # Get overall sizes from CutSizes—but only those that have been used in passports.
        # We'll build a set of (size, extra) pairs from all PassportSize records.
        passport_sizes_set = set()
        for passport in passports:
            for ps in passport.passport_sizes.all():
                # Use ps.extra or '' if it's empty.
                passport_sizes_set.add((ps.size_quantity.size, ps.extra or ''))
        # Sort them: first by numeric value of size (if possible), then by extra.
        try:
            passport_sizes_display = sorted(passport_sizes_set, key=lambda x: (int(x[0]), x[1]))
        except ValueError:
            passport_sizes_display = sorted(passport_sizes_set, key=lambda x: (x[0], x[1]))

        # Prepare total quantities per size (if needed elsewhere)
        total_quantity_per_size = defaultdict(int)
        for cs in cut.cut_sizes.all():
            # Note: This sums by size only (ignoring extra).
            total_quantity_per_size[cs.size_quantity.size] += cs.size_quantity.quantity

        # Total layers from all passports.
        total_layers = sum(passport.layers for passport in passports if passport.layers)

        context.update({
            'passports': passports,
            'total_quantity_per_size': dict(total_quantity_per_size),
            'total_layers': total_layers,
            'passport_sizes_display': passport_sizes_display,  # New variable for overall passport sizes.
            'sidebar_type': 'technology'
        })
        return context
    
@login_required
@technologist_required
def assign_operations(request, passport_id):
    passport = get_object_or_404(Passport, pk=passport_id)
    model_operations = ModelOperation.objects.filter(
        model=passport.cut.order.model
    ).select_related('operation').order_by('order')
    operations = [model_op.operation for model_op in model_operations]
    size_quantities = PassportSize.objects.filter(passport=passport).order_by('size_quantity__size')

    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = json.loads(request.body)
        operation_id = data.get('operation_id')
        passport_size_id = data.get('passport_size_id')
        value = data.get('value')

        try:
            with transaction.atomic():
                passport_size = PassportSize.objects.get(id=passport_size_id)
                total_quantity = passport_size.quantity

                entries = value.split(',')
                for entry in entries:
                    if '(' in entry and ')' in entry:
                        employee_id_input, quantity = entry.split('(')
                        employee_id_input = employee_id_input.strip()
                        quantity = int(quantity.strip(' )'))
                    else:
                        employee_id_input = entry
                        quantity = total_quantity

                    employee_profile = UserProfile.objects.filter(employee_id=employee_id_input).first()
                    if not employee_profile:
                        continue

                    work, created = Work.objects.get_or_create(
                        operation_id=operation_id,
                        passport_size=passport_size,
                    )
                    AssignedWork.objects.create(
                        work=work,
                        employee=employee_profile,
                        quantity=quantity
                    )
                    
                return JsonResponse({'status': 'success'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    work_by_op_and_size = {}
    for assigned_work in AssignedWork.objects.filter(work__passport_size__passport=passport).select_related('employee', 'work__operation', 'work__passport_size'):
        # Key as a tuple of operation_id and passport_size_id
        key = (assigned_work.work.operation_id, assigned_work.work.passport_size_id)
        if key not in work_by_op_and_size:
            work_by_op_and_size[key] = [assigned_work]
        else:
            work_by_op_and_size[key].append(assigned_work)

    return render(request, 'technologist/passports/assign_operations.html', {
        'passport': passport,
        'operations': operations,
        'size_quantities': size_quantities,
        'work_by_op_and_size': work_by_op_and_size,
        'sidebar_type': 'technology'
    })

@login_required
@technologist_required
def assign_operations_by_cut(request, cut_id):
    cut = get_object_or_404(Cut, pk=cut_id)
    model_operations = ModelOperation.objects.filter(
        model=cut.order.model
    ).select_related('operation').order_by('order')
    operations = [model_op.operation for model_op in model_operations]

    # Get all passports for this cut, ordered by number (or creation order)
    passports = cut.passports.all().order_by('number')
    
    # For each passport, get its passport sizes ordered by size.
    passport_sizes_by_passport = {}
    for passport in passports:
        sizes = passport.passport_sizes.all().order_by('size_quantity__size')
        passport_sizes_by_passport[passport.id] = sizes

    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        import json
        from django.db import transaction
        data = json.loads(request.body)
        operation_id = data.get('operation_id')
        passport_size_id = data.get('passport_size_id')
        value = data.get('value')

        try:
            with transaction.atomic():
                passport_size = PassportSize.objects.get(id=passport_size_id)
                total_quantity = passport_size.quantity

                entries = value.split(',')
                for entry in entries:
                    if '(' in entry and ')' in entry:
                        employee_id_input, quantity = entry.split('(')
                        employee_id_input = employee_id_input.strip()
                        quantity = int(quantity.strip(' )'))
                    else:
                        employee_id_input = entry.strip()
                        quantity = total_quantity

                    employee_profile = UserProfile.objects.filter(
                        employee_id=employee_id_input,
                        type=UserProfile.EMPLOYEE,
                        
                    ).first()
                    if not employee_profile:
                        raise Exception(f"Employee not found.")

                    work, created = Work.objects.get_or_create(
                        operation_id=operation_id,
                        passport_size=passport_size,
                    )
                    AssignedWork.objects.create(
                        work=work,
                        employee=employee_profile,
                        quantity=quantity
                    )
                return JsonResponse({'status': 'success'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # Build a dictionary of already assigned works, keyed by (operation_id, passport_size_id)
    work_by_op_and_size = {}
    for aw in AssignedWork.objects.filter(work__passport_size__passport__in=passports).select_related('employee', 'work__operation', 'work__passport_size'):
        key = (aw.work.operation_id, aw.work.passport_size_id)
        work_by_op_and_size.setdefault(key, []).append(aw)

    return render(request, 'technologist/passports/assign_operations_by_cut.html', {
        'cut': cut,
        'operations': operations,
        'passports': passports,
        'passport_sizes_by_passport': passport_sizes_by_passport,
        'work_by_op_and_size': work_by_op_and_size,
        'sidebar_type': 'technology'
    })

@require_POST
@technologist_required
@login_required
def update_passport_quantity(request):
    """
    Updates the factual quantity for a given PassportSize and updates all 
    AssignedWork records related to that PassportSize.
    Adjusts the corresponding SizeQuantity.factual based on the delta.
    """
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"status": "error", "message": "Invalid JSON."}, status=400)

    passport_size_id = data.get("passport_size_id")
    new_quantity = data.get("new_quantity")

    if passport_size_id is None or new_quantity is None:
        return JsonResponse({"status": "error", "message": "Missing parameters."}, status=400)

    try:
        new_quantity = int(new_quantity)
    except ValueError:
        return JsonResponse({"status": "error", "message": "Invalid quantity."}, status=400)

    try:
        passport_size = PassportSize.objects.get(id=passport_size_id)
    except PassportSize.DoesNotExist:
        return JsonResponse({"status": "error", "message": "PassportSize not found."}, status=404)

    # Calculate the delta
    old_factual = passport_size.factual or 0
    delta = new_quantity - old_factual

    # Update PassportSize factual
    passport_size.factual = new_quantity
    passport_size.save()

    # Adjust SizeQuantity factual
    size_qty = passport_size.size_quantity
    if size_qty:
        size_qty.factual = (size_qty.factual or 0) + delta
        size_qty.save()
    # Update all AssignedWork records
    AssignedWork.objects.filter(work__passport_size=passport_size).update(quantity=new_quantity)

    return JsonResponse({"status": "success"})

@login_required
@technologist_required
@require_POST
def update_work(request):
    data = json.loads(request.body)
    assigned_work_id = data.get('work_id')
    value = data.get('value')

    try:
        current_assignment = AssignedWork.objects.get(id=assigned_work_id)
        work = Work.objects.get(id=current_assignment.work.id)

        if not value.strip():
            current_assignment.delete()
            remaining_assignments = AssignedWork.objects.filter(work=work).exists()
            if not remaining_assignments:
                work.delete()
            return JsonResponse({'status': 'success'})

        new_assignments_data = value.split(',')
        first = True  # Flag to track the first item

        for item in new_assignments_data:
            employee_id, quantity = item.split('(')
            employee_id = employee_id.strip()
            quantity = int(quantity.strip(' )'))

            employee_profile = UserProfile.objects.filter(
                employee_id=employee_id, type=UserProfile.EMPLOYEE,
                ).first()

            if not employee_profile:
                raise Exception(f"Employee not found.")

            # Handle the first item by updating existing assigned work
            if first:
                current_assignment.employee = employee_profile
                current_assignment.quantity = quantity
                current_assignment.save()
                first = False  # Update the flag after handling the first item
            else:
                # Create new assigned works for subsequent items
                AssignedWork.objects.create(
                    work=work,
                    employee=employee_profile,
                    quantity=quantity
                )

        return JsonResponse({'status': 'success'})

    except Work.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Work not found'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@login_required
@technologist_required
def update_work_success(request):
    work_id = request.POST.get('work_id')
    is_success = request.POST.get('is_success') == 'true'

    try:
        assigned_work = AssignedWork.objects.get(pk=work_id)
        assigned_work.is_success = is_success
        assigned_work.save()

        return JsonResponse({'status': 'success'})
    except AssignedWork.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Assigned work not found'}, status=404)

@method_decorator([login_required, technologist_required], name='dispatch')
class OperationListView(ListView):
    model = Operation
    template_name = 'technologist/operations/list.html'
    context_object_name = 'operations'

    def get_paginate_by(self, queryset):
        node_id = self.request.GET.get('node', None)
        if node_id:
            return None
        return 15

    def get_queryset(self):
        queryset = super().get_queryset().filter(is_archived=False).order_by('number')
        node_id = self.request.GET.get('node', None)
        if node_id:
            queryset = queryset.filter(node_id=node_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        nodes = Node.objects.filter(is_archived=False).order_by('name')
        context['nodes'] = nodes
        context['selected_node'] = self.request.GET.get('node', '')
        context['upload_form'] = UploadFileForm()
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class OperationCreateView(CreateView):
    model = Operation
    form_class = OperationForm
    template_name = 'technologist/operations/create.html'
    success_url = reverse_lazy('operation_create')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class OperationDetailView(DetailView):
    model = Operation
    template_name = 'technologist/operations/detail.html'
    context_object_name = 'operation'
    def get_context_data(self, **kwargs):
        context = super(OperationDetailView, self).get_context_data(**kwargs)
        operation = self.object
        models = Model.objects.filter(operations=operation)
        context['models'] = models
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class OperationUpdateView(UpdateView):
    model = Operation
    form_class = OperationForm
    template_name = 'technologist/operations/edit.html'
    success_url = reverse_lazy('operation_list')  # Update this to your desired success URL

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    
@method_decorator([login_required, technologist_required], name='dispatch')
class OperationDeleteView(DeleteView):
    model = Operation
    template_name = 'technologist/operations/delete.html'
    success_url = reverse_lazy('operation_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    
@method_decorator([login_required, technologist_required], name='dispatch')
class ArchivedOperationListView(ListView):
    model = Operation
    template_name = 'technologist/operations/list.html'
    context_object_name = 'operations'
    paginate_by = 10

    def get_queryset(self):
        return Operation.objects.filter(is_archived=True).order_by('number')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        context['archived'] = True
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class OperationArchiveView(UpdateView):
    model = Operation
    template_name = 'technologist/operations/delete.html'
    success_url = reverse_lazy('operation_list')

    def post(self, request, *args, **kwargs):
        operation = self.get_object()
        operation.is_archived = True
        operation.save()
        return HttpResponseRedirect(self.success_url)

@method_decorator([login_required, technologist_required], name='dispatch')
class OperationUnArchiveView(UpdateView):
    model = Operation
    template_name = 'technologist/operations/delete.html'
    success_url = reverse_lazy('archived_operation_list')

    def post(self, request, *args, **kwargs):
        operation = self.get_object()
        operation.is_archived = False
        operation.save()
        return HttpResponseRedirect(self.success_url)

@login_required
@technologist_required
@require_POST
def operation_upload(request):
    form = UploadFileForm(request.POST, request.FILES)
    if form.is_valid():
        excel_file = request.FILES['excel_file']
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                number, operation_name, node_name, equipment_name, time, price = row

                node, _ = Node.objects.get_or_create(name=node_name)
                equipment, _ = Equipment.objects.get_or_create(name=equipment_name)
                operation, created = Operation.objects.get_or_create(
                    number=number,
                    defaults={
                        'name': operation_name,
                        'equipment': equipment,
                        'node': node,
                        'preferred_completion_time': time,
                        'payment': price
                    }
                )
                if not created:
                    operation.name = operation_name
                    operation.equipment = equipment
                    operation.node = node
                    operation.preferred_completion_time = time
                    operation.payment = price
                    operation.save()
                print(operation)
        messages.success(request, 'Operations uploaded successfully.')
        return HttpResponseRedirect(reverse_lazy('operation_list'))
    else:
        messages.error(request, 'There was an error with the file upload.')
        return HttpResponseRedirect(reverse_lazy('operation_upload'))
    
@login_required
@technologist_required
def operation_download(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Operations"

    headers = ["№ПП (авто генерация)", "Тех-процесс", "Узел", "№ узла", "Оборудование", "Время", "Оплата"]
    sheet.append(headers)

    for operation in Operation.objects.filter(is_archived=False).order_by('number'):
        row = [
            operation.number,
            operation.name,
            operation.node.name if operation.node else "",
            operation.equipment.name if operation.equipment else "",
            operation.preferred_completion_time,
            operation.payment,
        ]
        sheet.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=operations.xlsx'
    workbook.save(response)
    return response
    
@method_decorator([login_required, technologist_required], name='dispatch')
class AssortmentListView(ListView):
    model = Assortment
    template_name = 'technologist/assortments/list.html'
    context_object_name = 'assortments'
    paginate_by = 10
    def get_queryset(self):
        return super().get_queryset().filter(is_archived=False).order_by('name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    
    
@method_decorator([login_required, technologist_required], name='dispatch')
class ArchivedAssortmentListView(ListView):
    model = Assortment
    template_name = 'technologist/assortments/list.html'
    context_object_name = 'assortments'
    paginate_by = 10
    def get_queryset(self):
        return super().get_queryset().filter(is_archived=True).order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    
@method_decorator([login_required, technologist_required], name='dispatch')
class AssortmentArchiveView(UpdateView):
    model = Assortment
    template_name = 'technologist/assortments/delete.html'
    success_url = reverse_lazy('assortment_list')
    
    def post(self, request, *args, **kwargs):
        assortment = self.get_object()
        assortment.is_archived = True
        assortment.save()
        return HttpResponseRedirect(self.success_url)


@method_decorator([login_required, technologist_required], name='dispatch')
class AssortmentUnArchiveView(UpdateView):
    model = Assortment
    success_url = reverse_lazy('assortment_list')

    def post(self, request, *args, **kwargs):
        assortment = self.get_object()
        assortment.is_archived = False
        assortment.save()
        return HttpResponseRedirect(self.success_url)
    
@method_decorator([login_required, technologist_required], name='dispatch')
class AssortmentCreateView(CreateView):
    model = Assortment
    form_class = AssortmentForm
    template_name = 'technologist/assortments/create.html'
    def get_success_url(self):
        return reverse('assortment_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context


@method_decorator([login_required, technologist_required], name='dispatch')
class AssortmentDetailView(DetailView):
    model = Assortment
    template_name = 'technologist/assortments/detail.html'
    context_object_name = 'assortment'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class AssortmentUpdateView(UpdateView):
    model = Assortment
    form_class = AssortmentForm
    template_name = 'technologist/assortments/edit.html'
    def get_success_url(self):
        return reverse('assortment_detail', kwargs={'pk': self.kwargs.get('pk')})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class AssortmentDeleteView(DeleteView):
    model = Assortment
    template_name = 'technologist/assortments/delete.html'
    success_url = reverse_lazy('assortment_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@require_POST
@login_required
def add_assortment_api(request):
    form = AssortmentForm(request.POST)
    if form.is_valid():
        assortment = form.save()
        data = {
            'success': True,
            'assortment_id': assortment.id,
            'assortment_name': assortment.name,
        }
        return JsonResponse(data)
    else:
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)

@method_decorator([login_required, technologist_required], name='dispatch')
class ModelListView(ListView):
    model = Model
    template_name = 'technologist/models/list.html'
    context_object_name = 'models'
    paginate_by = 10

    def get_queryset(self):
        queryset = Model.objects.filter(is_archived=False).order_by('name')
        assortment = self.request.GET.get('assortment', None)
        if assortment:
            queryset = queryset.filter(assortment_id=assortment)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        # Assuming you have an Assortment model and you want only non-archived ones
        context['assortments'] = Assortment.objects.filter(is_archived=False).order_by('name')
        context['selected_assortment'] = self.request.GET.get('assortment', '')
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class ArchivedModelListView(ListView):
    model = Model
    template_name = 'technologist/models/list.html'
    context_object_name = 'models'
    paginate_by = 10
    def get_queryset(self):
        return Model.objects.filter(is_archived=True).order_by('name')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class ModelArchiveView(UpdateView):
    model = Model
    template_name = 'technologist/models/delete.html'

    def post(self, request, *args, **kwargs):
        model = self.get_object()
        model.is_archived = True
        model.save()
        return HttpResponseRedirect(reverse_lazy('model_list'))


@method_decorator([login_required, technologist_required], name='dispatch')
class ModelUnArchiveView(UpdateView):
    model = Model
    template_name = 'technologist/models/delete.html'

    def post(self, request, *args, **kwargs):
        model = self.get_object()
        model.is_archived = False
        model.save()
        return HttpResponseRedirect(reverse_lazy('model_list'))
        
@xframe_options_exempt
@login_required
@technologist_required
def model_create(request, pk=None):
    copy_id = request.GET.get('copy')
    original = get_object_or_404(Model, pk=copy_id) if copy_id else None
    modal = request.GET.get('modal') == '1'

    if request.method == 'POST':
        form = ModelCustomForm(request.POST, request.FILES, instance=None, copy_id=copy_id)
        if form.is_valid():
            model_instance = form.save()
            if copy_id and original:
                for bom in original.bill_of_materials.all():
                    BillOfMaterials.objects.create(
                        model=model_instance,
                        item=bom.item,
                        quantity=bom.quantity
                    )
            # ✅ Decide where to redirect:
            if modal:
                return JsonResponse({
                    'success': True,
                    'model_id': model_instance.id,
                    'model_name': model_instance.name
                })
            else:
                return redirect('model_detail', pk=model_instance.id)
    else:
        form = ModelCustomForm(instance=(original if copy_id else None), copy_id=copy_id)
        form.fields['operations'].queryset = Operation.objects.select_related('node').all().order_by(
            'node__name', 'name'
        )

    operations_order_json = ""
    if copy_id:
        operations_order = list(
            ModelOperation.objects.filter(model=original).order_by('order').values_list('operation_id', flat=True)
        )
        operations_order_json = json.dumps(operations_order)

    template_name = 'technologist/models/edit.html' if original else 'technologist/models/create.html'
    context = {
        'form': form,
        'is_copying': bool(copy_id),
        'copy_model': original if copy_id else None,
        'operations_order_json': operations_order_json,
        'sidebar_type': 'technology',
        'modal': modal,
    }
    return render(request, template_name, context)

@method_decorator([login_required, technologist_required], name='dispatch')
class ModelDetailView(DetailView):
    model = Model
    template_name = 'technologist/models/detail.html'
    context_object_name = 'model'

    def get_context_data(self, **kwargs):
        context = super(ModelDetailView, self).get_context_data(**kwargs)
        model = context['model']
        context['ordered_operations'] = model.operations.all().order_by('modeloperation__order')
        context['sidebar_type'] = 'technology'
        return context

@login_required
@technologist_required
def model_edit(request, pk):
    model_instance = get_object_or_404(Model, pk=pk)
    
    if request.method == 'POST':
        form = ModelCustomForm(request.POST, request.FILES, instance=model_instance)
        if form.is_valid():
            form.save()
            return redirect('model_detail', pk=pk)
    else:
        form = ModelCustomForm(instance=model_instance)
        form.fields['operations'].queryset = Operation.objects.select_related('node').all().order_by('node__name', 'name')
    
    # Define operations_order_json regardless of the method
    operations_order = list(ModelOperation.objects.filter(model=model_instance).order_by('order').values_list('operation_id', flat=True))
    operations_order_json = json.dumps(operations_order)
    
    return render(request, 'technologist/models/edit.html', {
        'form': form,
        'model': model_instance,
        'operations_order_json': operations_order_json,
        'sidebar_type': 'technology'
    })

@method_decorator([login_required, technologist_required], name='dispatch')
class ModelDeleteView(DeleteView):
    model = Model
    template_name = 'technologist/models/delete.html'
    def get_success_url(self):
        return reverse('model_list')



@method_decorator([login_required, technologist_required], name='dispatch')
class NodeListVIew(ListView):
    model = Node
    template_name = 'technologist/nodes/list.html'
    context_object_name = 'nodes'
    paginate_by = 10
    
    def get_queryset(self):
        return Node.objects.filter(is_archived=False).order_by('name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context


@method_decorator([login_required, technologist_required], name='dispatch')
class ArchivedNodeListView(ListView):
    model = Node
    template_name = 'technologist/nodes/list.html'
    context_object_name = 'nodes'
    paginate_by = 10

    def get_queryset(self):
        return Node.objects.filter(is_archived=True).order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class NodeArchiveView(UpdateView):
    model = Node
    template_name = 'technologist/nodes/delete.html'
    success_url = reverse_lazy('node_list')

    def post(self, request, *args, **kwargs):
        node = self.get_object()
        node.is_archived = True
        node.save()
        return HttpResponseRedirect(self.success_url)

@method_decorator([login_required, technologist_required], name='dispatch')
class NodeUnArchiveView(UpdateView):
    model = Node
    template_name = 'technologist/nodes/delete.html'
    success_url = reverse_lazy('node_list')

    def post(self, request, *args, **kwargs):
        node = self.get_object()
        node.is_archived = False
        node.save()
        return HttpResponseRedirect(self.success_url)

@method_decorator([login_required, technologist_required], name='dispatch')
class NodeCreateView(CreateView):
    model = Node
    form_class = NodeForm
    template_name = 'technologist/nodes/create.html'
    success_url = reverse_lazy('node_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class NodeDetailView(DetailView):
    model = Node
    template_name = 'technologist/nodes/detail.html'
    context_object_name = 'node'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class NodeUpdateView(UpdateView):
    model = Node
    form_class = NodeForm
    template_name = 'technologist/nodes/edit.html'
    success_url = reverse_lazy('node_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class NodeDeleteView(DeleteView):
    model = Node
    template_name = 'technologist/nodes/delete.html'
    success_url = reverse_lazy('node_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@require_POST
@login_required
def add_node_api(request):
    form = NodeForm(request.POST)
    if form.is_valid():
        node = form.save()
        data = {
            'success': True,
            'node_id': node.id,
            'node_name': node.name,
        }
        return JsonResponse(data)
    else:
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)


@require_POST
@login_required
def add_equipment_api(request):
    form = EquipmentForm(request.POST)
    if form.is_valid():
        equipment = form.save()
        data = {
            'success': True,
            'equipment_id': equipment.id,
            'equipment_name': equipment.name,
        }
        return JsonResponse(data)
    else:
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)

@method_decorator([login_required, technologist_required], name='dispatch')
class EquipmentListView(ListView):
    model = Equipment
    template_name = 'technologist/equipment/list.html'
    context_object_name = 'equipment'
    paginate_by = 10

    def get_queryset(self):
        return Equipment.objects.filter(is_archived=False).order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class ArchivedEquipmentListView(ListView):
    model = Equipment
    template_name = 'technologist/equipment/list.html'
    context_object_name = 'equipment'
    paginate_by = 10

    def get_queryset(self):
        return Equipment.objects.filter(is_archived=True).order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    
@method_decorator([login_required, technologist_required], name='dispatch')
class EquipmentArchiveView(UpdateView):
    model = Equipment
    template_name = 'technologist/equipment/delete.html'
    success_url = reverse_lazy('equipment_list')

    def post(self, request, *args, **kwargs):
        equipment = self.get_object()
        equipment.is_archived = True
        equipment.save()
        return HttpResponseRedirect(self.success_url)
    
@method_decorator([login_required, technologist_required], name='dispatch')
class EquipmentUnArchiveView(UpdateView):
    model = Equipment
    template_name = 'technologist/equipment/delete.html'
    success_url = reverse_lazy('equipment_list')

    def post(self, request, *args, **kwargs):
        equipment = self.get_object()
        equipment.is_archived = False
        equipment.save()
        return HttpResponseRedirect(self.success_url)
    
@method_decorator([login_required, technologist_required], name='dispatch')
class EquipmentCreateView(CreateView):
    model = Equipment
    form_class = EquipmentForm
    template_name = 'technologist/equipment/create.html'
    success_url = reverse_lazy('equipment_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class EquipmentDetailView(DetailView):
    model = Equipment
    template_name = 'technologist/equipment/detail.html'
    context_object_name = 'equipment'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class EquipmentUpdateView(UpdateView):
    model = Equipment
    form_class = EquipmentForm
    template_name = 'technologist/equipment/edit.html'
    success_url = reverse_lazy('equipment_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class EquipmentDeleteView(DeleteView):
    model = Equipment
    template_name = 'technologist/equipment/delete.html'
    success_url = reverse_lazy('equipment_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    
@login_required
@technologist_required
def update_assortment_name(request, pk):
    if request.method == 'POST':
        data = json.loads(request.body)
        assortment = Assortment.objects.get(pk=pk)
        assortment.name = data['name']
        assortment.save()
        return JsonResponse({'status': 'success', 'message': 'Assortment name updated successfully'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@method_decorator([login_required, technologist_required], name='dispatch')
class ColorListView(ListView):
    model = Color
    template_name = 'technologist/color/list.html'
    context_object_name = 'color'
    paginate_by = 10

    def get_queryset(self):
        return Color.objects.filter(is_archived=False).order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class ArchivedColorListView(ListView):
    model = Color
    template_name = 'technologist/color/list.html'
    context_object_name = 'color'
    paginate_by = 10

    def get_queryset(self):
        return Color.objects.filter(is_archived=True).order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    
@method_decorator([login_required, technologist_required], name='dispatch')
class ColorArchiveView(UpdateView):
    model = Color
    template_name = 'technologist/color/delete.html'
    success_url = reverse_lazy('color_list')

    def post(self, request, *args, **kwargs):
        color = self.get_object()
        color.is_archived = True
        color.save()
        return HttpResponseRedirect(self.success_url)
    
@method_decorator([login_required, technologist_required], name='dispatch')
class ColorUnArchiveView(UpdateView):
    model = Color
    template_name = 'technologist/color/delete.html'
    success_url = reverse_lazy('color_list')

    def post(self, request, *args, **kwargs):
        color = self.get_object()
        color.is_archived = False
        color.save()
        return HttpResponseRedirect(self.success_url)
    
@method_decorator([login_required, technologist_required], name='dispatch')
class ColorCreateView(CreateView):
    model = Color
    form_class = ColorForm
    template_name = 'technologist/color/create.html'
    success_url = reverse_lazy('color_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class ColorDetailView(DetailView):
    model = Color
    template_name = 'technologist/color/detail.html'
    context_object_name = 'color'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class ColorUpdateView(UpdateView):
    model = Color
    form_class = ColorForm
    template_name = 'technologist/color/edit.html'
    success_url = reverse_lazy('color_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class ColorDeleteView(DeleteView):
    model = Color
    template_name = 'technologist/color/delete.html'
    success_url = reverse_lazy('color_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context



@method_decorator([login_required, technologist_required], name='dispatch')
class FabricsListView(ListView):
    model = Fabrics
    template_name = 'technologist/fabrics/list.html'
    context_object_name = 'fabrics'
    paginate_by = 10

    def get_queryset(self):
        return Fabrics.objects.filter(is_archived=False).order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class ArchivedFabricsListView(ListView):
    model = Fabrics
    template_name = 'technologist/fabrics/list.html'
    context_object_name = 'fabrics'
    paginate_by = 10

    def get_queryset(self):
        return Fabrics.objects.filter(is_archived=True).order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    
@method_decorator([login_required, technologist_required], name='dispatch')
class FabricsArchiveView(UpdateView):
    model = Fabrics
    template_name = 'technologist/fabrics/delete.html'
    success_url = reverse_lazy('fabrics_list')

    def post(self, request, *args, **kwargs):
        fabric = self.get_object()
        fabric.is_archived = True
        fabric.save()
        return HttpResponseRedirect(self.success_url)
    
@method_decorator([login_required, technologist_required], name='dispatch')
class FabricsUnArchiveView(UpdateView):
    model = Fabrics
    template_name = 'technologist/fabrics/delete.html'
    success_url = reverse_lazy('fabrics_list')

    def post(self, request, *args, **kwargs):
        fabric = self.get_object()
        fabric.is_archived = False
        fabric.save()
        return HttpResponseRedirect(self.success_url)
    
@method_decorator([login_required, technologist_required], name='dispatch')
class FabricsCreateView(CreateView):
    model = Fabrics
    form_class = FabricsForm
    template_name = 'technologist/fabrics/create.html'
    success_url = reverse_lazy('fabrics_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class FabricsDetailView(DetailView):
    model = Fabrics
    template_name = 'technologist/fabrics/detail.html'
    context_object_name = 'fabrics'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class FabricsUpdateView(UpdateView):
    model = Fabrics
    form_class = FabricsForm
    template_name = 'technologist/fabrics/edit.html'
    success_url = reverse_lazy('fabrics_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class FabricsDeleteView(DeleteView):
    model = Fabrics
    template_name = 'technologist/fabrics/delete.html'
    success_url = reverse_lazy('fabrics_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    

@login_required
@technologist_required
def bom_create(request, pk):
    size_qty = get_object_or_404(SizeQuantity, pk=pk)
    order = size_qty.orders.first()

    # All size quantities for navigation
    all_sqs = list(order.size_quantities.all().order_by('id'))
    idx = all_sqs.index(size_qty)

    # Build ordered (color_id, fabrics_id) group sequence following the same order as all_sqs
    group_seq = []
    group_to_first_sq = {}  # (color_id, fabrics_id) -> first SQ (for landing)
    for sq in all_sqs:
        key = (sq.color_id, sq.fabrics_id)
        if key not in group_seq:
            group_seq.append(key)
            group_to_first_sq[key] = sq

    # Current group key
    cur_key = (size_qty.color_id, size_qty.fabrics_id)
    cur_group_index = group_seq.index(cur_key)

    if request.method == 'POST':
        # ---------- parse scope ----------
        apply_scope = request.POST.get('apply_scope', 'this')  # 'this' | 'same_cf' | 'all_order'

        # ---------- Parse posted rows once ----------
        rows = []
        row = 0
        while f'row-{row}_item' in request.POST:
            item_id = request.POST.get(f'row-{row}_item')
            qty_str  = request.POST.get(f'row-{row}_quantity')
            if item_id and qty_str:
                item = get_object_or_404(Item, pk=item_id)
                quantity = Decimal(qty_str)
                rows.append((item, quantity))
            row += 1

        # If nothing was entered, just bounce back
        if not rows:
            return redirect('bom_create', pk=size_qty.pk)

        # ---------- Determine target SizeQuantity set ----------
        if apply_scope == 'same_cf':
            # same color + fabric as current, across this order
            target_sqs = order.size_quantities.filter(
                color_id=size_qty.color_id,
                fabrics_id=size_qty.fabrics_id,
            ).order_by('id')
        elif apply_scope == 'all_order':
            # all sizes in this order (color/fabric agnostic)
            target_sqs = order.size_quantities.all().order_by('id')
        else:
            # only this size (current behavior)
            target_sqs = SizeQuantity.objects.filter(pk=size_qty.pk)

        # ---------- REMOVED: Stock placeholder creation ----------

        # ---------- Apply rows to all targets atomically ----------
        with transaction.atomic():
            for sq in target_sqs:
                sq.bill_of_materials.all().delete()
                for item, quantity in rows:
                    BillOfMaterials.objects.create(
                        sizequantity=sq,
                        item=item,
                        quantity=quantity
                    )

        # ---------- Navigation after write ----------
        if apply_scope == 'this':
            # Step to next size within all_sqs; if none, go to order detail
            if idx + 1 < len(all_sqs):
                return redirect('bom_create', pk=all_sqs[idx + 1].pk)
            return redirect('order_detail', pk=order.pk)

        if apply_scope == 'same_cf':
            # Step to next (color,fabric) group; if none, go to order detail
            next_group_index = cur_group_index + 1
            if next_group_index < len(group_seq):
                next_key = group_seq[next_group_index]
                next_sq = group_to_first_sq[next_key]  # first SQ of next group
                return redirect('bom_create', pk=next_sq.pk)
            return redirect('order_detail', pk=order.pk)

        # apply_scope == 'all_order' → done; go to order detail
        return redirect('order_detail', pk=order.pk)

    # ---------- GET (unchanged except context additions) ----------
    categories = Category.objects.filter(is_archived=False).order_by('name')
    items = Item.objects.filter(is_archived=False).order_by('name')
    items_by_cat = defaultdict(list)
    for it in items:
        items_by_cat[it.category_id].append(it)

    boms_qs = size_qty.bill_of_materials.all()
    if not boms_qs.exists() and idx > 0:
        boms_qs = all_sqs[idx - 1].bill_of_materials.all()

    colors = Color.objects.filter(is_archived=False)
    fabrics = Fabrics.objects.filter(is_archived=False)
    suppliers = Supplier.objects.filter(is_archived=False)
    order_client = order.client_order.client if order and order.client_order else None

    return render(request, 'technologist/models/bom_create.html', {
        'order': order,
        'sizequantity': size_qty,
        'categories': categories,
        'items_by_category': dict(items_by_cat),
        'boms': boms_qs,
        'sidebar_type': 'technology',
        'colors': colors,
        'fabrics': fabrics,
        'suppliers': suppliers,
        'order_client': order_client,
        'cf_hint': f"{getattr(size_qty.color,'name','')}/{getattr(size_qty.fabrics,'name','')}",
    })

@require_POST
@login_required
def add_category_api(request):
    name = request.POST.get('name', '').strip()
    is_fabric = request.POST.get('is_fabric') == 'on'

    if not name:
        return JsonResponse(
            {'success': False, 'errors': {'name': ['This field is required.']}},
            status=400
        )
    cat = Category.objects.create(name=name, is_fabric=is_fabric)
    return JsonResponse({
        'success': True,
        'category_id': cat.pk,
        'category_name': cat.name,
        'is_fabric': cat.is_fabric
    })


@login_required
def items_by_category_api(request):
    cat_id = request.GET.get('category')
    if not cat_id:
        return JsonResponse({'items': []})
    qs = Item.objects.filter(category_id=cat_id, is_archived=False).order_by('name')
    items = [{'id': i.pk, 'name': i.name, 'unit': i.unit} for i in qs]
    return JsonResponse({'items': items})


@require_POST
@login_required
def add_item_api(request):
    form = ItemForm(request.POST)
    if form.is_valid():
        item = form.save()
        data = {
            'success': True,
            'item_id': item.id,
            'item_name': item.name,
            'unit': item.unit
        }
        return JsonResponse(data)
    else:
        # Return form errors as JSON (status code 400)
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    
@require_POST
@login_required
def add_fabric_item_api(request):
    form = FabricItemForm(request.POST)

    if form.is_valid():
        fabric_item = form.save(commit=False)  # Don't save yet
        
        # Auto-generate name and unit
        color_name = fabric_item.color.name if fabric_item.color else ''
        fabric_name = fabric_item.fabric.name if fabric_item.fabric else ''
        supplier_name = fabric_item.supplier.name if fabric_item.supplier else ''
        client_name = fabric_item.client.name if fabric_item.client else ''
        width_value = f"{fabric_item.width:.2f}" if fabric_item.width else ''

        fabric_item.name = f"{color_name} {fabric_name} {width_value}м для {client_name} от {supplier_name}".strip()
        fabric_item.unit = "м"

        fabric_item.save()  # Now save
        
        return JsonResponse({
            'success': True,
            'item_id': fabric_item.pk,
            'item_name': fabric_item.name,
        })
    else:
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)

@method_decorator([login_required, technologist_required], name='dispatch')
class OrderBomView(DetailView):
    model = Order  # switched from SizeQuantity to Order
    template_name = 'technologist/orders/bom.html'
    context_object_name = 'order'  # now the object is the order itself

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        order = self.object

        # days left
        today = timezone.now().date()
        term = order.client_order.term
        context['days_left'] = max((term - today).days, 0)

        # bring in all size‐quantities + their BOM entries
        context['size_quantities'] = (
            order.size_quantities
                 .prefetch_related('bill_of_materials__item__category')
                 .all()
        )
        context['first_sq'] = context['size_quantities'].first().pk if context['size_quantities'] else None
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class EmployeeListTechnologistView(ListView):
    model = UserProfile
    template_name = 'technologist/employees/list.html'
    context_object_name = 'employees'
    paginate_by = 10

    def get_queryset(self):
        return UserProfile.objects.filter(
            is_archived=False
        ).exclude(type=UserProfile.ADMIN).annotate(
            employee_id_int=Cast('employee_id', IntegerField())
        ).order_by('employee_id_int')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['upload_form'] = UploadFileForm()
        context['sidebar_type'] = 'technology'
        
        return context
    
@method_decorator([login_required, technologist_required], name='dispatch')
class EmployeeCreateTechnologistView(CreateView):
    template_name = 'technologist/employees/create.html'
    form_class = UserWithProfileTechnologistForm
    success_url = reverse_lazy('employee_list_technologist')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@method_decorator([login_required, technologist_required], name='dispatch')
class EmployeeDetailTechnologistView(DetailView):
    model = UserProfile
    template_name = 'technologist/employees/detail.html'
    context_object_name = 'employee'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

@login_required
@technologist_required
def employee_edit_technologist(request, pk):
    user_profile = get_object_or_404(UserProfile, pk=pk)
    user = user_profile.user

    if request.method == 'POST':
        user_form = UserEditForm(request.POST, instance=user)
        if user_form.is_valid():
            user_form.save()
            messages.success(request, 'Employee details updated successfully.')
            return redirect('employee_list_technologist')
    else:
        user_form = UserEditTechnologistForm(instance=user)

    context = {'user_form': user_form, 'user_profile': user_profile}
    context['sidebar_type'] = 'technology'
    return render(request, 'technologist/employees/edit.html', context)

@method_decorator([login_required, technologist_required], name='dispatch')
class EmployeeDeleteTechnologistView(DeleteView):
    model = UserProfile
    template_name = 'technologist/employees/delete.html'
    success_url = reverse_lazy('employee_list_technologist')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    
@method_decorator([login_required, technologist_required], name='dispatch')
class EmployeeArchiveTechnologistView(UpdateView):
    model = UserProfile
    template_name = 'technologist/employees/list.html'
    success_url = reverse_lazy('employee_list_technologist')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    
    def post(self, request, *args, **kwargs):
        employee = self.get_object()
        employee.is_archived = True
        employee.save()
        return HttpResponseRedirect(self.success_url)
   
@method_decorator([login_required, technologist_required], name='dispatch')
class EmployeeUnArchiveTechnologistView(UpdateView):
    model = UserProfile
    template_name = 'technologist/employees/list.html'
    success_url = reverse_lazy('employee_list_technologist')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context
    
    def post(self, request, *args, **kwargs):
        employee = self.get_object()
        employee.is_archived = False
        employee.save()
        return HttpResponseRedirect(self.success_url)
     
@method_decorator([login_required, technologist_required], name='dispatch')
class ArchivedEmployeeListTechnologistView(ListView):
    template_name = 'technologist/employees/list.html'
    context_object_name = 'employees'
    paginate_by = 10
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sidebar_type'] = 'technology'
        return context

    def get_queryset(self):
        return UserProfile.objects.filter(
            is_archived=True
            ).order_by('employee_id')
    

@login_required
@technologist_required
@require_POST
def employee_upload_technologist(request):
    form = UploadFileForm(request.POST, request.FILES)
    if form.is_valid():
        excel_file = request.FILES['excel_file']
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            # Get the current company from thread-local storage.
            current_company = get_current_company()
            if current_company is None:
                messages.error(request, 'No company found in context.')
                return redirect(reverse_lazy('employee_list_technologist'))

            with transaction.atomic():
                # Iterate rows skipping the header (starting at row 2)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    first_name, last_name, employee_id, emp_type, password = row
                    
                    # Generate username using current company ID and employee ID.
                    username = f"{current_company.id}-{employee_id}"
                    
                    try:
                        # Because of your custom manager, this lookup is already company-aware.
                        profile = UserProfile.objects.get(employee_id=employee_id)
                    except UserProfile.DoesNotExist:
                        # Create a new user and UserProfile.
                        user = User.objects.create(
                            username=username,
                            first_name=first_name,
                            last_name=last_name
                        )
                        user.set_password(password)
                        user.save()
                        # The save method on CompanyAwareModel will assign the current company.
                        profile = UserProfile.objects.create(
                            user=user,
                            employee_id=employee_id,
                            type=int(emp_type) if emp_type is not None else UserProfile.EMPLOYEE
                        )
                    else:
                        # Update existing user and profile.
                        user = profile.user
                        user.first_name = first_name
                        user.last_name = last_name
                        user.username = username
                        user.set_password(password)
                        user.save()
                        
                        profile.type = int(emp_type) if emp_type is not None else UserProfile.EMPLOYEE
                        profile.save()

            messages.success(request, 'Employees uploaded successfully.')
            return redirect(reverse_lazy('employee_list_technologist'))
        except Exception as e:
            messages.error(request, f'Error processing the file: {e}')
            return redirect(reverse_lazy('employee_list_technologist'))
        finally:
            workbook.close()
    else:
        messages.error(request, 'Invalid file format.')
        return redirect(reverse_lazy('employee_list_technologist'))
